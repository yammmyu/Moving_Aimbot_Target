import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import date, timedelta

F = "Arial"
NAVY = "1A3A5C"; LTBLUE = "DCE6EF"; BAND = "F2F6F9"; GOLD = "FDF6E7"
thin = Side(style="thin", color="B8C6D2")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

def hdr(ws, row, labels, widths=None, height=30):
    for i, lab in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font = Font(name=F, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = BORD
    ws.row_dimensions[row].height = height
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)

def body(ws, r0, rows, wrap_cols=(), bold_cols=()):
    for ri, row in enumerate(rows):
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=r0 + ri, column=ci, value=val)
            c.font = Font(name=F, size=10, bold=(ci in bold_cols))
            c.alignment = Alignment(vertical="top", wrap_text=(ci in wrap_cols))
            c.border = BORD
            if ri % 2:
                c.fill = PatternFill("solid", fgColor=BAND)

def title(ws, text, sub=None, span=6):
    ws["A1"] = text
    ws["A1"].font = Font(name=F, bold=True, size=14, color=NAVY)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name=F, size=9, italic=True, color="666666")
    ws.row_dimensions[1].height = 20

# ============================================================ 1. ROADMAP
ws = wb.active
ws.title = "Roadmap"
title(ws, "Aim-Truth Rig — 16-Week Phased Roadmap",
      "Week 1 = w/c 2026-08-03. Effort in solo-days. Update the Status column; % complete rolls up on the Dashboard.")

HDRS = ["Wk", "Week of", "Phase", "Work item", "Deliverable / exit criterion",
        "Depends on", "Effort (days)", "Status", "Risk link"]
hdr(ws, 4, HDRS, widths=[5, 11, 26, 40, 52, 12, 11, 13, 10], height=34)

W1 = date(2026, 8, 3)
R = [
    (1, "P0 Observability", "Archaeology spike on inherited stack (timebox 4 days)", "Written one-pager: detector, filter type + rate, latency scheme, gimbal protocol. Explicit decision: instrument vs. replace tracker.", "—", 4, "R2"),
    (1, "P0 Observability", "Confirm camera model + trigger/strobe capability (Q1)", "Datasheet line confirming external trigger in OR strobe out. GO/NO-GO on full spin range.", "—", 0.5, "R1"),
    (1, "P0 Observability", "Confirm compute headroom + disk throughput for full-rate logging (Q4)", "Measured write bandwidth vs. required frame+state rate. No dropped frames under load.", "—", 0.5, "R2"),
    (2, "P0 Observability", "Structured estimator-state logging (REQ-0.1)", "Every filter update logged, not just final aim output. Self-describing dataset format w/ metadata.", "Wk1 spike", 3, "—"),
    (2, "P0 Observability", "Gimbal commanded-vs-actual angle logging (REQ-0.2)", "Second channel on same time base. Enables estimator-vs-servo error split.", "REQ-0.1", 1, "—"),
    (2, "P0 Observability", "Order long-lead parts (motor if not owned, MCU, frame)", "Parts ordered. Piezo parts deliberately NOT ordered yet.", "Q1 answered", 0.5, "R5"),
    (3, "P0 Observability", "Hardware frame trigger / strobe capture (REQ-0.3)", "MCU drives or captures camera exposure; encoder latched in same ISR.", "Parts", 3, "R1"),
    (3, "P0 Observability", "LED sync verification test", "Measured sync error written into dataset metadata. MUST be <= 200 us.", "REQ-0.3", 1.5, "R1"),
    (3, "GATE", "WEEK 3 HARD GATE — sync error <= 200 us?", "If > 1 ms with no hardware path: descope to static-target check, abandon spinning rig.", "LED test", 0, "R1"),
    (4, "P1 Estimator", "Build spin rig: frame, hub, plate mount (REQ-1.1)", "Direct-drive GM6020 mount. Static / constant-speed / step-change modes. Speed held to 2%.", "Wk3 gate", 4, "R7"),
    (4, "P1 Estimator", "Safety: guard, hardware E-stop, captive fasteners", "E-stop cuts motor power in hardware, not software. Verified by test.", "Frame", 1, "SAFETY"),
    (5, "P1 Estimator", "Encoder truth pipeline at >=1 kHz, <=0.05 deg (REQ-1.2)", "Absolute plate angle on rig time base. Validated against high-fps phone video.", "Rig built", 2, "R8"),
    (5, "P1 Estimator", "Analysis: bias / latency / noise decomposition (REQ-1.3)", "Signed error vs. time + 3 scalars per spin rate. Latency via cross-correlation lag.", "Encoder truth", 3, "—"),
    (6, "P1 Estimator", "Estimator-vs-servo error split (REQ-1.4)", "Same analysis on gimbal feedback channel. Error budget attributed.", "REQ-1.3", 1.5, "—"),
    (6, "P1 Estimator", "Baseline report at 0.5 / 1.5 / 3.0 rev/s", "One-page report. This is the Week 6 baseline that M5 is measured against.", "REQ-1.4", 1.5, "—"),
    (6, "GATE", "WEEK 6 PAYBACK GATE — did rig data change a decision?", "If no tuning decision was made on rig data: PARK until off-season.", "Baseline", 0, "R3"),
    (7, "P4 Replay", "Deterministic replay harness (REQ-4.1)", "Dataset in + params in -> error metrics out. Bit-for-bit reproducible across runs.", "Wk6 gate", 4, "—"),
    (7, "P2 Live-fire", "Confirm firing space + referee set availability (Q2, Q3)", "Booked space at fixed range; referee set confirmed or Phase 2 goes paper-only.", "—", 0.5, "R4"),
    (8, "P4 Replay", "Parameter sweep driver + scalar cost function (REQ-4.2)", "Grid/random search ranking configs. Runs in seconds, no hardware.", "REQ-4.1", 2.5, "—"),
    (8, "P2 Live-fire", "NOISE FLOOR: static plate, clamped launcher, witness paper (REQ-2.1)", "Group spread of the launcher alone. Hard floor on achievable accuracy. Do this BEFORE any hit-rate number.", "Space", 1.5, "—"),
    (8, "P2 Live-fire", "Per-shot muzzle velocity logging (REQ-2.2)", "Velocity joined to time-of-flight and required lead angle.", "Referee link", 1, "R4"),
    (9, "P2 Live-fire", "Hit rate + group offset vs. spin rate (REQ-2.3)", "Chart with dispersion floor drawn as horizontal reference. Witness paper digitised per magazine.", "REQ-2.1", 3, "—"),
    (9, "P4 Replay", "Held-out validation split (REQ-4.3)", "Tune on set A, report on set B. Guards against dataset memorisation.", "REQ-4.2", 1.5, "R6"),
    (10, "P3 Piezo", "GO/NO-GO BENCH TEST: 4 piezos, spare plate, grid of taps (REQ-3.1)", "Calibrated RMS error on a known grid. NO-GO if > 25 mm -> reallocate to P4.", "Wk9", 3, "R5"),
    (10, "GATE", "WEEK 10 PHASE-3 GATE + M4 adoption check", "Piezo <= 25 mm RMS? AND >= 3 tuning decisions made on rig data?", "Bench test", 0, "R5"),
    (11, "P3 Piezo", "Comparator front end + STM32 input-capture timing", "4-channel threshold crossings timestamped on rig time base. ns-class resolution.", "Wk10 gate", 3, "—"),
    (12, "P3 Piezo", "Multilateration + empirical calibration map (REQ-3.2)", "Grid-derived map, NOT an analytical wave-speed model. Stored with rig config.", "Front end", 4, "—"),
    (13, "P3 Piezo", "Integrate onto live plate; join impacts to estimator state (REQ-3.3)", "Each impact carries rig timestamp; paired with estimator state at firing minus TOF.", "Calibration", 3, "—"),
    (13, "P4 Replay", "Real-robot validation of swept parameters (REQ-4.4)", "No offline-selected parameter adopted without live validation.", "REQ-4.3", 2, "R6"),
    (14, "P3 Piezo", "Accuracy acceptance: <=25 mm RMS, >=95% quadrant (M7)", "Measured on held-out grid positions, not calibration positions.", "Integration", 2, "R5"),
    (14, "P1 Estimator", "M5 re-measurement vs. Week 6 baseline", "Same rig, same protocol. Target: >= 30% estimator error reduction.", "P4 tuning", 1.5, "—"),
    (15, "Consolidate", "Regenerate every reported number from stored datasets", "Single command per figure. No number that exists only in a screenshot.", "All", 2, "—"),
    (15, "Consolidate", "Dataset archive + metadata audit", "Every recording labelled with rig config, sync error, camera settings, git commit.", "All", 1.5, "—"),
    (16, "Consolidate", "Final report: error budget, tuning outcomes, known limits", "Includes an honest section on what the rig does NOT measure (see PRD R6).", "Wk15", 2.5, "R6"),
    (16, "Consolidate", "Off-season backlog: what v2 should do differently", "Written while the pain is fresh.", "—", 1, "—"),
]

r0 = 5
rows = []
for wk, phase, item, deliv, dep, eff, risk in R:
    rows.append([wk, W1 + timedelta(weeks=wk - 1), phase, item, deliv, dep, eff, "Not started", risk])
body(ws, r0, rows, wrap_cols=(3, 4, 5, 6), bold_cols=(1,))

last = r0 + len(rows) - 1
for r in range(r0, last + 1):
    ws.cell(row=r, column=2).number_format = "dd-mmm"
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=r, column=7).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=r, column=8).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=r, column=9).alignment = Alignment(horizontal="center", vertical="top")
    if ws.cell(row=r, column=3).value == "GATE":
        for c in range(1, 10):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FCE4D6")
            ws.cell(row=r, column=c).font = Font(name=F, size=10, bold=True, color="843C0C")

dv = DataValidation(type="list", formula1='"Not started,In progress,Blocked,Done,Descoped"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"H{r0}:H{last}")
ws.conditional_formatting.add(f"H{r0}:H{last}", CellIsRule(operator="equal", formula=['"Done"'], fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(name=F, size=10, color="006100")))
ws.conditional_formatting.add(f"H{r0}:H{last}", CellIsRule(operator="equal", formula=['"Blocked"'], fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(name=F, size=10, color="9C0006")))
ws.conditional_formatting.add(f"H{r0}:H{last}", CellIsRule(operator="equal", formula=['"In progress"'], fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(name=F, size=10, color="9C6500")))

tot = last + 2
ws.cell(row=tot, column=6, value="TOTAL EFFORT (solo-days)").font = Font(name=F, bold=True, size=10)
tc = ws.cell(row=tot, column=7, value=f"=SUM(G{r0}:G{last})")
tc.font = Font(name=F, bold=True, size=10); tc.alignment = Alignment(horizontal="center")
ws.cell(row=tot + 1, column=6, value="Available (16 wks x 2.5 days/wk realistic solo, in-season)").font = Font(name=F, size=9, italic=True)
ac = ws.cell(row=tot + 1, column=7, value=40); ac.font = Font(name=F, size=10, bold=True, color="0000FF"); ac.alignment = Alignment(horizontal="center")
ws.cell(row=tot + 2, column=6, value="Headroom (negative = overcommitted)").font = Font(name=F, bold=True, size=10)
hc = ws.cell(row=tot + 2, column=7, value=f"=G{tot+1}-G{tot}")
hc.font = Font(name=F, bold=True, size=10); hc.alignment = Alignment(horizontal="center")
ws.cell(row=tot + 3, column=6, value="Blue = your input. Adjust available days to your real capacity.").font = Font(name=F, size=9, italic=True, color="666666")
ws.cell(row=tot + 5, column=1, value="CAPACITY REALITY CHECK: the plan as scoped is ~66.5 solo-days against ~40 available at 2.5 days/week. Dropping Phase 3 entirely (15 days) still leaves ~51.5.").font = Font(name=F, size=9, italic=True, bold=True, color="9C0006")
ws.cell(row=tot + 6, column=1, value="Closing the gap needs one of: (a) ~3.3 days/week instead of 2.5, (b) drop Phase 3 AND cut Phase 2 to a single dispersion-floor session, or (c) extend to 20 weeks.").font = Font(name=F, size=9, italic=True, color="9C0006")
ws.cell(row=tot + 7, column=1, value="This is deliberate. A plan that fits perfectly on day one has not been costed honestly. Decide which lever to pull at the Week 6 payback gate, when there is evidence.").font = Font(name=F, size=9, italic=True, color="666666")
ws.sheet_view.showGridLines = False

# ============================================================ 2. DASHBOARD
ds = wb.create_sheet("Dashboard")
title(ds, "Dashboard", "All figures are formulas over the Roadmap and Risks tabs — they update as you edit Status.")
ds.column_dimensions["A"].width = 44
for col in "BCDE":
    ds.column_dimensions[col].width = 15

ds["A4"] = "PROGRESS"; ds["A4"].font = Font(name=F, bold=True, size=11, color=NAVY)
prog = [
    ("Total work items", f"=COUNTA(Roadmap!D{r0}:D{last})-COUNTIF(Roadmap!C{r0}:C{last},\"GATE\")"),
    ("Done", f"=COUNTIF(Roadmap!H{r0}:H{last},\"Done\")"),
    ("In progress", f"=COUNTIF(Roadmap!H{r0}:H{last},\"In progress\")"),
    ("Blocked", f"=COUNTIF(Roadmap!H{r0}:H{last},\"Blocked\")"),
    ("Descoped", f"=COUNTIF(Roadmap!H{r0}:H{last},\"Descoped\")"),
    ("% complete", f"=IFERROR(B6/B5,0)"),
]
for i, (lab, f_) in enumerate(prog):
    r = 5 + i
    ds.cell(row=r, column=1, value=lab).font = Font(name=F, size=10, bold=(lab == "% complete"))
    c = ds.cell(row=r, column=2, value=f_)
    c.font = Font(name=F, size=10, bold=(lab == "% complete")); c.alignment = Alignment(horizontal="center"); c.border = BORD
ds["B10"].number_format = "0%"

ds["A13"] = "EFFORT BY PHASE (solo-days)"; ds["A13"].font = Font(name=F, bold=True, size=11, color=NAVY)
hdr(ds, 14, ["Phase", "Planned", "Done", "% done"], widths=None, height=18)
ds.freeze_panes = None
phases = ["P0 Observability", "P1 Estimator", "P2 Live-fire", "P3 Piezo", "P4 Replay", "Consolidate"]
for i, ph in enumerate(phases):
    r = 15 + i
    ds.cell(row=r, column=1, value=ph).font = Font(name=F, size=10)
    for col, f_ in ((2, f'=SUMIF(Roadmap!$C${r0}:$C${last},$A{r},Roadmap!$G${r0}:$G${last})'),
                    (3, f'=SUMIFS(Roadmap!$G${r0}:$G${last},Roadmap!$C${r0}:$C${last},$A{r},Roadmap!$H${r0}:$H${last},"Done")'),
                    (4, f'=IFERROR(C{r}/B{r},0)')):
        c = ds.cell(row=r, column=col, value=f_)
        c.font = Font(name=F, size=10); c.alignment = Alignment(horizontal="center"); c.border = BORD
    ds.cell(row=r, column=1).border = BORD
    ds.cell(row=r, column=4).number_format = "0%"
r = 21
ds.cell(row=r, column=1, value="TOTAL").font = Font(name=F, size=10, bold=True)
for col, f_ in ((2, "=SUM(B15:B20)"), (3, "=SUM(C15:C20)"), (4, "=IFERROR(C21/B21,0)")):
    c = ds.cell(row=r, column=col, value=f_); c.font = Font(name=F, size=10, bold=True); c.alignment = Alignment(horizontal="center"); c.border = BORD
ds.cell(row=r, column=1).border = BORD
ds["D21"].number_format = "0%"

ds["A24"] = "GATES — the decisions that keep this project honest"
ds["A24"].font = Font(name=F, bold=True, size=11, color=NAVY)
hdr(ds, 25, ["Gate", "Week", "Test", "Action on failure"], widths=None, height=18)
gates = [
    ("Sync gate", 3, "Measured camera-to-encoder sync error <= 200 us", "If > 1 ms with no hardware path: descope to static-target check, abandon spinning rig"),
    ("Payback gate", 6, "Has rig data changed at least one tuning decision?", "Park the project until off-season; robot work wins during the season"),
    ("Phase 3 gate", 10, "Piezo bench RMS <= 25 mm on a known grid", "Do not integrate. Reallocate the 4 weeks to Phase 4 replay tuning"),
    ("Adoption gate", 10, ">= 3 tuning decisions made on rig data (M4)", "The tool is not being used. Stop building; find out why"),
    ("Drift tripwire", "any", "2 consecutive weeks with no dataset recorded", "You are building infrastructure for its own sake. Record something crude, re-plan"),
]
for i, g in enumerate(gates):
    r = 26 + i
    for ci, v in enumerate(g, start=1):
        c = ds.cell(row=r, column=ci, value=v)
        c.font = Font(name=F, size=10); c.alignment = Alignment(vertical="top", wrap_text=True); c.border = BORD
        if i % 2: c.fill = PatternFill("solid", fgColor=BAND)
    ds.row_dimensions[r].height = 30
ds.column_dimensions["C"].width = 46
ds.column_dimensions["D"].width = 58
ds.freeze_panes = None
ds.sheet_view.showGridLines = False

# ============================================================ 3. BOM
bs = wb.create_sheet("BOM")
title(bs, "Bill of Materials", "Blue = your input (adjust qty/price to what you actually pay or already own). Budget ceiling from kickoff: $500.")
hdr(bs, 4, ["Item", "Purpose", "Phase", "Qty", "Unit $", "Line $", "Already own?", "Order by"],
    widths=[30, 38, 8, 6, 9, 10, 13, 11], height=30)
bom = [
    ("GM6020 gimbal motor", "Direct-drive spin + absolute encoder ground truth (0.044 deg, no backlash)", "P1", 1, 110, "Likely", "Wk 2"),
    ("STM32F4 dev board", "Rig time base, CAN, camera trigger, piezo input capture", "P0", 1, 15, "Maybe", "Wk 2"),
    ("USB-CAN adapter", "Motor command + host link", "P1", 1, 25, "Likely", "Wk 2"),
    ("24 V PSU", "Motor supply (or reuse a team battery)", "P1", 1, 25, "Maybe", "Wk 2"),
    ("Aluminium extrusion + base plate + fasteners", "Rigid frame; must not walk under vibration (risk R7)", "P1", 1, 60, "No", "Wk 2"),
    ("3D-printed hub + sensor mounts", "Interface to standard armor plate; captive fasteners", "P1", 1, 8, "In-house", "Wk 3"),
    ("Opto-isolator + wiring + connectors", "Camera trigger isolation", "P0", 1, 12, "No", "Wk 2"),
    ("LED + driver", "Sync verification test (REQ-0.3)", "P0", 1, 8, "No", "Wk 2"),
    ("Carbon / witness paper", "Phase 2 spatial ground truth. Best value per dollar in this BOM.", "P2", 1, 5, "No", "Wk 7"),
    ("Piezo discs", "Impact localisation elements", "P3", 4, 2, "No", "Wk 10*"),
    ("Comparator + analogue front end parts", "Threshold crossing detection for TDOA", "P3", 1, 27, "No", "Wk 10*"),
    ("Safety guard / polycarbonate shield", "Spinning plate containment + eye protection zone", "P1", 1, 30, "No", "Wk 4"),
]
b0 = 5
for i, (item, purpose, ph, qty, unit, own, when) in enumerate(bom):
    r = b0 + i
    vals = [item, purpose, ph, qty, unit, f"=D{r}*E{r}", own, when]
    for ci, v in enumerate(vals, start=1):
        c = bs.cell(row=r, column=ci, value=v)
        c.font = Font(name=F, size=10, color=("0000FF" if ci in (4, 5) else "000000"))
        c.alignment = Alignment(vertical="top", wrap_text=(ci == 2), horizontal=("center" if ci in (3, 4, 7, 8) else "general"))
        c.border = BORD
        if i % 2: c.fill = PatternFill("solid", fgColor=BAND)
    bs.cell(row=r, column=5).number_format = '$#,##0'
    bs.cell(row=r, column=6).number_format = '$#,##0'
bl = b0 + len(bom) - 1
r = bl + 2
bs.cell(row=r, column=5, value="SUBTOTAL (all phases)").font = Font(name=F, bold=True, size=10)
c = bs.cell(row=r, column=6, value=f"=SUM(F{b0}:F{bl})"); c.font = Font(name=F, bold=True, size=10); c.number_format = '$#,##0'; c.border = BORD
bs.cell(row=r+1, column=5, value="Less items likely already owned").font = Font(name=F, size=10)
c = bs.cell(row=r+1, column=6, value=f'=SUMIF(G{b0}:G{bl},"Likely",F{b0}:F{bl})'); c.font = Font(name=F, size=10); c.number_format = '$#,##0'; c.border = BORD
bs.cell(row=r+2, column=5, value="Less Phase 3 (gated at Wk 10)").font = Font(name=F, size=10)
c = bs.cell(row=r+2, column=6, value=f'=SUMIF(C{b0}:C{bl},"P3",F{b0}:F{bl})'); c.font = Font(name=F, size=10); c.number_format = '$#,##0'; c.border = BORD
bs.cell(row=r+3, column=5, value="COMMITTED SPEND BEFORE WK 10").font = Font(name=F, bold=True, size=10)
c = bs.cell(row=r+3, column=6, value=f"=F{r}-F{r+1}-F{r+2}"); c.font = Font(name=F, bold=True, size=10); c.number_format = '$#,##0'; c.border = BORD
bs.cell(row=r+5, column=5, value="Budget ceiling").font = Font(name=F, size=10)
c = bs.cell(row=r+5, column=6, value=500); c.font = Font(name=F, size=10, color="0000FF"); c.number_format = '$#,##0'; c.border = BORD
bs.cell(row=r+6, column=5, value="Headroom vs. worst case (buy everything)").font = Font(name=F, bold=True, size=10)
c = bs.cell(row=r+6, column=6, value=f"=F{r+5}-F{r}"); c.font = Font(name=F, bold=True, size=10); c.number_format = '$#,##0'; c.border = BORD
bs.cell(row=r+8, column=1, value="* Do not order Phase 3 parts until the Week 10 go/no-go bench test passes. Ordering early converts a reversible decision into a sunk cost.").font = Font(name=F, size=9, italic=True, color="843C0C")
bs.cell(row=r+9, column=1, value="Assumption: GM6020 and USB-CAN adapter marked 'Likely' owned — confirm against the team's spares inventory in Week 1.").font = Font(name=F, size=9, italic=True, color="666666")
bs.sheet_view.showGridLines = False

# ============================================================ 4. RISKS
rs = wb.create_sheet("Risks")
title(rs, "Risk Register", "Score = Impact x Likelihood (1-5 each). Review at every gate. Blue cells are your scoring inputs.")
hdr(rs, 4, ["ID", "Risk", "Impact\n(1-5)", "Likelihood\n(1-5)", "Score", "Severity", "Mitigation", "Owner", "Review at"],
    widths=[6, 40, 9, 11, 8, 11, 56, 8, 12], height=34)
risks = [
    ("R1", "Time sync cannot reach +/-200 us (camera lacks trigger/strobe line)", 5, 3, "Confirm camera trigger capability in Week 1 before any other purchase. Fallback: cap spin rate at ~1 rev/s so the artefact stays below the signal, and state the reduced range in every report.", "Self", "Wk 3 gate"),
    ("R2", "Inherited stack undocumented, unbuildable, or unloggable", 4, 4, "Timeboxed 4-day archaeology spike ending in a written decision: instrument it, or replace the tracker with a minimal known-good one. Do not drift into an open-ended refactor.", "Self", "Wk 1"),
    ("R3", "Solo builder + live season -> project starved of time", 4, 4, "Every phase gate delivers standalone value. Week 6 payback gate. Prefer a finished Phase 1 over a half-finished Phase 3.", "Self", "Wk 6 gate"),
    ("R4", "No spare referee system available for the target rig", 2, 3, "Witness paper covers spatial measurement independently, so referee counting is optional not blocking. Confirm availability by Week 6 (Q2).", "Self", "Wk 7"),
    ("R5", "Piezo localisation misses the 25 mm accuracy target", 3, 3, "Week 10 bench go/no-go on a spare plate before any integration or full parts order. Reallocate to Phase 4 on a no-go.", "Self", "Wk 10 gate"),
    ("R6", "Overfitting to the rig: tuned params help on the bench, hurt on a real robot", 5, 4, "The rig measures a proxy, not the goal. Mandate real-robot validation before adopting any tuned parameter (REQ-4.4). Keep held-out datasets (REQ-4.3). Vary lighting, range and background between recordings.", "Self", "Wk 13"),
    ("R7", "Frame resonance or walk corrupts encoder truth at high spin", 3, 2, "Sweep spin rate looking for anomalous encoder residuals. Ballast or clamp the frame. Record a static-plate control run every session.", "Self", "Wk 5"),
    ("R8", "Rig ground truth trusted without independent validation", 4, 3, "Film the plate with a phone at high frame rate and confirm the encoder-reported rate matches. Cheap, and it catches whole classes of wiring and unit errors.", "Self", "Wk 5"),
    ("R9", "Injury from spinning plate or projectile", 5, 2, "Hardware E-stop cutting motor power (not software). Mechanical guard. Captive fasteners on the hub. Eye protection whenever the launcher is armed. Non-negotiable.", "Self", "Wk 4"),
    ("R10", "Logging drops frames under load, biasing datasets toward easy scenes", 4, 3, "Measure disk throughput and CPU headroom in Week 1 (Q4). Assert on dropped-frame counters and fail loudly rather than silently.", "Self", "Wk 2"),
]
q0 = 5
for i, (rid, risk, imp, lik, mit, own, rev) in enumerate(risks):
    r = q0 + i
    vals = [rid, risk, imp, lik, f"=C{r}*D{r}",
            f'=IF(E{r}>=15,"Critical",IF(E{r}>=9,"High",IF(E{r}>=4,"Medium","Low")))', mit, own, rev]
    for ci, v in enumerate(vals, start=1):
        c = rs.cell(row=r, column=ci, value=v)
        c.font = Font(name=F, size=10, bold=(ci == 1), color=("0000FF" if ci in (3, 4) else "000000"))
        c.alignment = Alignment(vertical="top", wrap_text=(ci in (2, 7)),
                                horizontal=("center" if ci in (1, 3, 4, 5, 6, 8, 9) else "general"))
        c.border = BORD
        if i % 2: c.fill = PatternFill("solid", fgColor=BAND)
rl = q0 + len(risks) - 1
rs.conditional_formatting.add(f"F{q0}:F{rl}", CellIsRule(operator="equal", formula=['"Critical"'], fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(name=F, size=10, bold=True, color="9C0006")))
rs.conditional_formatting.add(f"F{q0}:F{rl}", CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill("solid", fgColor="FFD9CC"), font=Font(name=F, size=10, bold=True, color="C55A11")))
rs.conditional_formatting.add(f"F{q0}:F{rl}", CellIsRule(operator="equal", formula=['"Medium"'], fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(name=F, size=10, color="9C6500")))
rs.conditional_formatting.add(f"F{q0}:F{rl}", CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(name=F, size=10, color="006100")))
rs.cell(row=rl + 2, column=2, value="Highest score:").font = Font(name=F, bold=True, size=10)
c = rs.cell(row=rl + 2, column=5, value=f"=MAX(E{q0}:E{rl})"); c.font = Font(name=F, bold=True, size=10); c.alignment = Alignment(horizontal="center")
rs.cell(row=rl + 3, column=2, value="Severity thresholds: Critical >=15, High >=9, Medium >=4, Low <4.").font = Font(name=F, size=9, italic=True, color="666666")
rs.sheet_view.showGridLines = False

# ============================================================ 5. METRICS
ms = wb.create_sheet("Metrics")
title(ms, "Success Metrics", "Blue cells are yours to fill in as measurements land. Status computes automatically where a target is numeric.")
hdr(ms, 4, ["ID", "Metric", "Target", "Actual", "Due", "Status", "How measured"],
    widths=[6, 34, 30, 14, 9, 13, 48], height=30)
mets = [
    ("M1", "Ground-truth uncertainty at 3 rev/s", "<= 0.25 deg", None, "Wk 3", "LED sync test + encoder spec, recorded in dataset metadata"),
    ("M2", "Estimator error characterised", "Bias, latency, noise at 3 spin rates", None, "Wk 6", "Phase 1 report; latency by cross-correlation lag"),
    ("M3", "Tuning cycle time", "<= 10 min, param change to number", None, "Wk 8", "Stopwatch over 5 trials"),
    ("M4", "Real adoption", ">= 3 tuning decisions on rig data", None, "Wk 10", "Decision log. THE metric that matters most."),
    ("M5", "Measured improvement", ">= 30% error reduction vs. Wk 6 baseline", None, "Wk 14", "Same rig, same protocol, re-measured"),
    ("M6", "Live-fire baseline", "Hit rate vs. spin rate, with dispersion floor", None, "Wk 9", "Phase 2 chart from witness paper + referee counts"),
    ("M7", "Impact localisation (if P3 proceeds)", "<= 25 mm RMS, >= 95% quadrant", None, "Wk 14", "Held-out grid positions, not calibration positions"),
]
m0 = 5
for i, (mid, met, tgt, act, due, how) in enumerate(mets):
    r = m0 + i
    vals = [mid, met, tgt, act, due, "Pending", how]
    for ci, v in enumerate(vals, start=1):
        c = ms.cell(row=r, column=ci, value=v)
        c.font = Font(name=F, size=10, bold=(ci == 1), color=("0000FF" if ci == 4 else "000000"))
        c.alignment = Alignment(vertical="top", wrap_text=(ci in (2, 3, 7)),
                                horizontal=("center" if ci in (1, 4, 5, 6) else "general"))
        c.border = BORD
        if i % 2: c.fill = PatternFill("solid", fgColor=BAND)
ml = m0 + len(mets) - 1
dv2 = DataValidation(type="list", formula1='"Pending,On track,At risk,Met,Missed,Descoped"', allow_blank=True)
ms.add_data_validation(dv2); dv2.add(f"F{m0}:F{ml}")
ms.conditional_formatting.add(f"F{m0}:F{ml}", CellIsRule(operator="equal", formula=['"Met"'], fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(name=F, size=10, color="006100")))
ms.conditional_formatting.add(f"F{m0}:F{ml}", CellIsRule(operator="equal", formula=['"Missed"'], fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(name=F, size=10, color="9C0006")))
ms.conditional_formatting.add(f"F{m0}:F{ml}", CellIsRule(operator="equal", formula=['"At risk"'], fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(name=F, size=10, color="9C6500")))
ms.sheet_view.showGridLines = False

# ============================================================ 6. OPEN QUESTIONS
qs = wb.create_sheet("Open Questions")
title(qs, "Open Questions", "Q1 and Q4 gate Week 1 and should be answered before anything is ordered.")
hdr(qs, 4, ["ID", "Question", "Blocks", "Needed by", "Status", "Answer"],
    widths=[6, 50, 38, 11, 13, 34], height=30)
qq = [
    ("Q1", "What camera does the aimbot use, and does it expose an external trigger input or a strobe output?", "REQ-0.3, the whole time-base design, and the usable spin-rate range", "Wk 1"),
    ("Q4", "What compute runs the aimbot, and does it have spare CPU and disk throughput for full-rate logging?", "REQ-0.1 - logging that drops frames under load produces biased datasets", "Wk 1"),
    ("Q5", "Which armor plate size is the reference target: small (~135 mm) or large (~235 mm)?", "Hub design, piezo geometry, calibration grid", "Wk 2"),
    ("Q2", "Is a spare referee system set available to mount on the target rig?", "Phase 2 hit counting (not spatial measurement, which paper covers)", "Wk 6"),
    ("Q3", "Is there a space where the launcher can be fired safely and repeatedly at a fixed range?", "All of Phase 2; may force a shorter test range", "Wk 6"),
    ("Q6", "Will anyone else on the team use this, even informally?", "Whether any documentation effort is justified at all", "Wk 8"),
]
p0 = 5
for i, (qid, q, blk, due) in enumerate(qq):
    r = p0 + i
    vals = [qid, q, blk, due, "Open", ""]
    for ci, v in enumerate(vals, start=1):
        c = qs.cell(row=r, column=ci, value=v)
        c.font = Font(name=F, size=10, bold=(ci == 1), color=("0000FF" if ci == 6 else "000000"))
        c.alignment = Alignment(vertical="top", wrap_text=(ci in (2, 3, 6)), horizontal=("center" if ci in (1, 4, 5) else "general"))
        c.border = BORD
        if i % 2: c.fill = PatternFill("solid", fgColor=BAND)
    qs.row_dimensions[r].height = 32
ql = p0 + len(qq) - 1
dv3 = DataValidation(type="list", formula1='"Open,Answered,Blocked,Moot"', allow_blank=True)
qs.add_data_validation(dv3); dv3.add(f"E{p0}:E{ql}")
qs.conditional_formatting.add(f"E{p0}:E{ql}", CellIsRule(operator="equal", formula=['"Answered"'], fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(name=F, size=10, color="006100")))
qs.conditional_formatting.add(f"E{p0}:E{ql}", CellIsRule(operator="equal", formula=['"Open"'], fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(name=F, size=10, color="9C6500")))
qs.cell(row=ql + 2, column=2, value="Legend: blue cells are for you to fill in. Answer Q1 before ordering anything - it determines whether the project's core premise holds.").font = Font(name=F, size=9, italic=True, color="666666")
qs.sheet_view.showGridLines = False

# ============================================================ 7. SYNC BUDGET
sb = wb.create_sheet("Sync Budget")
title(sb, "Time-Base Error Budget", "The calculation behind the project's hardest requirement. Blue cells are inputs - change them to test your own setup.")
sb.column_dimensions["A"].width = 42
for col in "BCDE":
    sb.column_dimensions[col].width = 16

sb["A4"] = "INPUTS"; sb["A4"].font = Font(name=F, bold=True, size=11, color=NAVY)
inputs = [("Max target spin rate (rev/s)", 3.0, "0.0"),
          ("Assumed clock skew, software timestamps (ms)", 2.0, "0.0"),
          ("Assumed clock skew, hardware trigger (ms)", 0.05, "0.000"),
          ("Encoder counts per revolution (GM6020)", 8192, "#,##0"),
          ("Smallest aiming error we want to resolve (deg)", 0.5, "0.00")]
for i, (lab, val, fmt) in enumerate(inputs):
    r = 5 + i
    sb.cell(row=r, column=1, value=lab).font = Font(name=F, size=10)
    c = sb.cell(row=r, column=2, value=val)
    c.font = Font(name=F, size=10, bold=True, color="0000FF")
    c.fill = PatternFill("solid", fgColor="FFFFCC")
    c.alignment = Alignment(horizontal="center"); c.border = BORD; c.number_format = fmt

sb["A12"] = "DERIVED"; sb["A12"].font = Font(name=F, bold=True, size=11, color=NAVY)
derived = [
    ("Angular velocity at max spin (deg/s)", "=B5*360", "#,##0"),
    ("Apparent error from software timestamps (deg)", "=B13*(B6/1000)", "0.000"),
    ("Apparent error from hardware trigger (deg)", "=B13*(B7/1000)", "0.000"),
    ("Encoder quantisation (deg)", "=360/B8", "0.000"),
    ("Combined hardware-path uncertainty (deg, RSS)", "=SQRT(B15^2+B16^2)", "0.000"),
    ("Software path as % of the signal we want to resolve", "=B14/B9", "0%"),
    ("Hardware path as % of the signal we want to resolve", "=B17/B9", "0%"),
]
for i, (lab, f_, fmt) in enumerate(derived):
    r = 13 + i
    sb.cell(row=r, column=1, value=lab).font = Font(name=F, size=10, bold=(i >= 4))
    c = sb.cell(row=r, column=2, value=f_)
    c.font = Font(name=F, size=10, bold=(i >= 4)); c.alignment = Alignment(horizontal="center"); c.border = BORD; c.number_format = fmt

sb["A21"] = "VERDICT"; sb["A21"].font = Font(name=F, bold=True, size=11, color=NAVY)
sb["A22"] = '=IF(B17<=0.25,"PASS - hardware-triggered path meets REQ-T2 (<=0.25 deg).","FAIL - reduce spin rate or improve triggering.")'
sb["A22"].font = Font(name=F, size=10, bold=True)
sb["A23"] = '=IF(B18>0.5,"Software timestamping fabricates more than half the error you are trying to measure. Do not use it.","Software timestamping may be tolerable at this spin rate.")'
sb["A23"].font = Font(name=F, size=10, italic=True, color="9C0006")

sb["A26"] = "SENSITIVITY: apparent angular error (deg) by spin rate and clock skew"
sb["A26"].font = Font(name=F, bold=True, size=11, color=NAVY)
skews = [0.05, 0.2, 1.0, 2.0, 5.0]
sb.cell(row=27, column=1, value="Spin rate (rev/s)  \\  Clock skew (ms)").font = Font(name=F, bold=True, size=9)
sb.cell(row=27, column=1).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
sb.cell(row=27, column=1).fill = PatternFill("solid", fgColor=NAVY)
sb.cell(row=27, column=1).font = Font(name=F, bold=True, size=9, color="FFFFFF")
sb.cell(row=27, column=1).border = BORD
for j, sk in enumerate(skews):
    c = sb.cell(row=27, column=2 + j, value=sk)
    c.font = Font(name=F, bold=True, size=10, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center"); c.border = BORD; c.number_format = "0.00"
rates = [0.5, 1.0, 1.5, 2.0, 3.0]
for i, rt in enumerate(rates):
    r = 28 + i
    c = sb.cell(row=r, column=1, value=rt)
    c.font = Font(name=F, bold=True, size=10); c.alignment = Alignment(horizontal="center"); c.border = BORD; c.number_format = "0.0"
    c.fill = PatternFill("solid", fgColor=LTBLUE)
    for j in range(len(skews)):
        cl = get_column_letter(2 + j)
        cc = sb.cell(row=r, column=2 + j, value=f"=$A{r}*360*({cl}$27/1000)")
        cc.font = Font(name=F, size=10); cc.alignment = Alignment(horizontal="center"); cc.border = BORD; cc.number_format = "0.00"
sb.conditional_formatting.add("B28:F32", CellIsRule(operator="greaterThan", formula=["0.5"], fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(name=F, size=10, color="9C0006")))
sb.conditional_formatting.add("B28:F32", CellIsRule(operator="between", formula=["0.25", "0.5"], fill=PatternFill("solid", fgColor="FFEB9C")))
sb.conditional_formatting.add("B28:F32", CellIsRule(operator="lessThan", formula=["0.25"], fill=PatternFill("solid", fgColor="C6EFCE")))
sb.cell(row=34, column=1, value="Red = the measurement artefact exceeds the 0.5 deg error you are trying to detect. Green = comfortably below it.").font = Font(name=F, size=9, italic=True, color="666666")
sb.cell(row=35, column=1, value="Source of encoder spec: DJI GM6020 absolute encoder, 8192 counts/rev. All other values are inputs you can change.").font = Font(name=F, size=9, italic=True, color="666666")
sb.sheet_view.showGridLines = False

wb.calculation.fullCalcOnLoad = True
wb.save("/sessions/great-vigilant-edison/mnt/outputs/Aim-Truth-Rig-Roadmap.xlsx")
print("saved")
