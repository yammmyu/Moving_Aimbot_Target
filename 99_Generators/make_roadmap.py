import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import date, timedelta

F = "Arial"
NAVY = "1A3A5C"; BAND = "F2F6F9"; LTBLUE = "DCE6EF"
thin = Side(style="thin", color="B8C6D2")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
BLUE = "0000FF"

wb = openpyxl.Workbook()

def hdr(ws, row, labels, widths=None, height=30, freeze=True):
    for i, lab in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font = Font(name=F, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = BORD
    ws.row_dimensions[row].height = height
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=1)

def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = Font(name=F, bold=True, size=14, color=NAVY)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name=F, size=9, italic=True, color="666666")
    ws.row_dimensions[1].height = 20

def putrow(ws, r, vals, wrap=(), center=(), blue=(), bold=(), band=False):
    for ci, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=ci, value=v)
        c.font = Font(name=F, size=10, bold=(ci in bold), color=(BLUE if ci in blue else "000000"))
        c.alignment = Alignment(vertical="top", wrap_text=(ci in wrap),
                                horizontal=("center" if ci in center else "general"))
        c.border = BORD
        if band:
            c.fill = PatternFill("solid", fgColor=BAND)

# ============================================================ ROADMAP
ws = wb.active
ws.title = "Roadmap"
title(ws, "Portable Aimbot Target - 26-Week Roadmap",
      "Week 1 = w/c 2026-08-03. Phase 1 ships a usable workshop target by Week 6. Update Status; the Dashboard rolls up.")
hdr(ws, 4, ["Wk", "Week of", "Phase", "Work item", "Deliverable / exit criterion", "Depends on", "Days", "Status", "Risk"],
    widths=[5, 11, 22, 42, 52, 13, 7, 13, 8], height=34)

W1 = date(2026, 8, 3)
R = [
 (1,"1 Workshop target","Freeze head-to-base interface spec (REQ-M6)","Bolt/slot pattern, vertical datum, single 24V+data connector. Written down and treated as frozen.","-",2,"-"),
 (1,"1 Workshop target","Define transport envelope + weight budget baseline","Case shape candidates costed; every head part gets a max-dimension and mass allocation before it is designed.","-",1.5,"R3"),
 (1,"1 Workshop target","Confirm fabrication access (Q7)","3D printer / machining / hand-cut only. Sets realistic effort for every mechanical part.","-",0.5,"-"),
 (2,"1 Workshop target","Head hub + arm design, inertia budgeted (REQ-M1, M2)","Head inertia estimated BEFORE build; plate radius minimised (inertia scales with r^2).","Interface",4,"R9"),
 (2,"1 Workshop target","Workshop base - deliberately crude, heavy, ballasted (REQ-M7)","Steel plate and ballast. No design effort spent here on purpose.","Interface",1,"-"),
 (3,"1 Workshop target","Fabricate head; route wiring through GM6020 hollow shaft (REQ-M4)","Head assembled. 18mm hollow shaft carries power/data on the rotation axis.","Design",4,"-"),
 (3,"1 Workshop target","Joint design pass: locate vs. clamp on every joint (REQ-M10..M13)","Every structural joint has a locating feature AND a separate preload element. All fasteners captive.","Head",2,"R1"),
 (4,"1 Workshop target","Motor control: 0.25-3.0 rev/s, speed hold within 2% (REQ-M2)","Commanded vs. actual logged. Static hold mode works. Spin-up time measured.","Head",3,"R9"),
 (4,"1 Workshop target","Safety: hardware E-stop, balance verification, guarding (R11)","E-stop cuts motor power in HARDWARE. Balance checked before first full-speed run.","Motor ctrl",1.5,"R10"),
 (5,"1 Workshop target","Endurance session + full joint re-check (R1)","Run at full speed for a realistic session, then inspect every joint. Loose joints found NOW, not in month 5.","Safety",1.5,"R1"),
 (5,"1 Workshop target","Weigh every head part against the budget","Actual masses replace estimates in the Weight Budget tab.","Head",0.5,"R3"),
 (6,"1 Workshop target","Phase 1 acceptance: a TEAMMATE runs a full session unaided (M2)","Someone who did not build it sets up and practises without asking how. Tests the design, not your memory.","Endurance",1,"R5"),
 (6,"GATE","WEEK 6 ADOPTION GATE - does a teammate reach for it unprompted?","If not, understand why before adding rail, sensors or a suitcase to it.","Session",0,"R5"),
 (5,"2 Rail","Rail spec: length, mounting, workshop constraints (Q5)","Confirmed against the actual space. Belt drive, stationary motor at one end (REQ-M8).","-",1,"-"),
 (7,"2 Rail","Build rail: extrusion, V-wheel carriage, belt drive","Carriage runs smoothly full length. Limit switches fitted.","Rail spec",4,"-"),
 (8,"2 Rail","Carriage mounts head via the frozen interface","Head bolts on with no rework. Proves the interface spec was right.","Rail, Interface",1,"-"),
 (9,"2 Rail","Unified control: translation + spin together","One interface commands both axes. Repeatable motion patterns.","Carriage",3,"-"),
 (10,"2 Rail","Rail acceptance: reliable over a full session","No jams, no belt skip, no limit-switch faults across a practice session.","Control",1,"-"),
 (8,"3 Plates + UI","ONE prototype instrumented plate (REQ-E5)","Single plate, 4 piezos, comparator front end. Built to learn, not to ship.","Wk6 gate",4,"R4"),
 (9,"3 Plates + UI","Bench characterisation against a tap grid (REQ-E4)","Empirical calibration map. Measured RMS error on HELD-OUT grid positions.","Prototype",3,"R4"),
 (10,"GATE","WEEK 10 PLATE GATE - RMS <= 25mm on the prototype?","If no: do not build four. Ship with stock plates + referee counting, which is still a good training tool.","Bench test",0,"R4"),
 (11,"3 Plates + UI","Rotating-frame electronics: head MCU + slip ring (REQ-E1..E3)","Only digital events cross the interface. Slip ring carries 24V + digital, ~4 conductors not 16+.","Wk10 gate",4,"R11"),
 (12,"3 Plates + UI","Build remaining 3 plates; per-plate calibration","All four calibrated and characterised. Replaceable impact faces fitted (REQ-M14).","Electronics",4,"R4"),
 (13,"3 Plates + UI","Boot self-test: 16 channels, motor, encoder, referee (REQ-E6)","Power-up health check. This is what makes the machine trustworthy.","Plates",2.5,"R4"),
 (14,"3 Plates + UI","Web UI: Live + Health screens","Spin rate, speed control, live impacts on 4 plate outlines, green/red health per item.","Self-test",4,"-"),
 (15,"3 Plates + UI","Web UI: Session + History + CSV export","Hit distribution, group centre and spread, session comparison.","UI core",3,"-"),
 (16,"3 Plates + UI","Referee cross-check in the workshop (REQ-E7)","Custom plate hit count vs. referee count. Agreement validates both.","UI",2,"R12"),
 (17,"3 Plates + UI","Phase 3 acceptance: 25mm RMS all plates, UI usable unaided","A teammate uses the UI without explanation.","All P3",1.5,"-"),
 (14,"4 Travel kit","Travel stand concept + packing study","Stand design against the FROZEN head geometry. Longest irreducible part identified.","Head final",3,"R6"),
 (16,"4 Travel kit","Choose case shape from the largest irreducible part (S7.1)","Case proportions selected AFTER the part list, not before. Least wasted volume.","Packing study",1,"R3"),
 (18,"4 Travel kit","Build travel stand: tool-free, packs flat, on-site ballast (REQ-M9)","Assembles without tools. Stable at 3 rev/s on a smooth floor.","Concept",5,"R6"),
 (20,"4 Travel kit","Custom case + cut foam, pocket per part (REQ-M14, M15)","Empty pocket = visible missing part. Assembly order printed inside the lid.","Stand",4,"R8"),
 (21,"GATE","WEEK 20-21 PACKING TEST - under 23kg and 158cm?","If not, decide deliberately: oversize fee, freight, or workshop-only. Not at the check-in desk.","Case",0,"R3"),
 (22,"4 Travel kit","Drop test representative of baggage handling (REQ-M16)","Head and plates survive. Failures found at home, not in China.","Case",1.5,"R8"),
 (23,"4 Travel kit","Spares kit defined and packed inside the case (REQ-M17)","Piezo modules, spare plate face, latches, fasteners. Fixed part of the packout.","Drop test",1,"R8"),
 (24,"4 Travel kit","TIMED DRY RUN: teammate assembles from case, unaided (M7)","Under 10 minutes, first attempt, no instructions from you. Tests the design, not your memory.","Spares",1.5,"R6"),
 (25,"4 Travel kit","Fix everything the dry run exposed","Second timed run passes.","Dry run",3,"-"),
 (26,"4 Travel kit","Travel documentation: packout card, customs list, setup card","One page. Lives in the case.","Fixes",1.5,"-"),
]
r0 = 5
for i, (wk, ph, item, deliv, dep, eff, risk) in enumerate(R):
    putrow(ws, r0 + i, [wk, W1 + timedelta(weeks=wk - 1), ph, item, deliv, dep, eff, "Not started", risk],
           wrap=(3, 4, 5, 6), center=(1, 2, 7, 8, 9), bold=(1,), band=(i % 2 == 1))
last = r0 + len(R) - 1
for r in range(r0, last + 1):
    ws.cell(row=r, column=2).number_format = "dd-mmm"
    if ws.cell(row=r, column=3).value == "GATE":
        for c in range(1, 10):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FCE4D6")
            ws.cell(row=r, column=c).font = Font(name=F, size=10, bold=True, color="843C0C")

dv = DataValidation(type="list", formula1='"Not started,In progress,Blocked,Done,Descoped"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"H{r0}:H{last}")
for val, fill, col in (("Done","C6EFCE","006100"), ("Blocked","FFC7CE","9C0006"), ("In progress","FFEB9C","9C6500")):
    ws.conditional_formatting.add(f"H{r0}:H{last}", CellIsRule(operator="equal", formula=[f'"{val}"'],
        fill=PatternFill("solid", fgColor=fill), font=Font(name=F, size=10, color=col)))

t = last + 2
ws.cell(row=t, column=6, value="TOTAL EFFORT (solo-days)").font = Font(name=F, bold=True, size=10)
c = ws.cell(row=t, column=7, value=f"=SUM(G{r0}:G{last})"); c.font = Font(name=F, bold=True, size=10); c.alignment = Alignment(horizontal="center")
ws.cell(row=t+1, column=6, value="Available (26 wks x 2.5 days/wk, in-season)").font = Font(name=F, size=9, italic=True)
c = ws.cell(row=t+1, column=7, value=65); c.font = Font(name=F, bold=True, size=10, color=BLUE); c.alignment = Alignment(horizontal="center")
ws.cell(row=t+2, column=6, value="Headroom (negative = overcommitted)").font = Font(name=F, bold=True, size=10)
c = ws.cell(row=t+2, column=7, value=f"=G{t+1}-G{t}"); c.font = Font(name=F, bold=True, size=10); c.alignment = Alignment(horizontal="center")
ws.cell(row=t+4, column=1, value="Phase 1 alone is ~22 days and delivers the training tool that removes the robot dependency. Every later phase is independently useful.").font = Font(name=F, size=9, italic=True, color="666666")
ws.cell(row=t+5, column=1, value="If headroom is negative, the intended sacrifice order is: Phase 4 polish, then Phase 3 UI depth, then the rail. Phase 1 is never sacrificed.").font = Font(name=F, size=9, italic=True, color="9C0006")
ws.sheet_view.showGridLines = False

# ============================================================ DASHBOARD
ds = wb.create_sheet("Dashboard")
title(ds, "Dashboard", "Formulas over the other tabs - updates as you edit Status.")
ds.column_dimensions["A"].width = 46
for col in "BCDE": ds.column_dimensions[col].width = 15

ds["A4"] = "PROGRESS"; ds["A4"].font = Font(name=F, bold=True, size=11, color=NAVY)
for i, (lab, f_) in enumerate([
    ("Total work items", f'=COUNTA(Roadmap!D{r0}:D{last})-COUNTIF(Roadmap!C{r0}:C{last},"GATE")'),
    ("Done", f'=COUNTIF(Roadmap!H{r0}:H{last},"Done")'),
    ("In progress", f'=COUNTIF(Roadmap!H{r0}:H{last},"In progress")'),
    ("Blocked", f'=COUNTIF(Roadmap!H{r0}:H{last},"Blocked")'),
    ("% complete", "=IFERROR(B6/B5,0)")]):
    r = 5 + i
    ds.cell(row=r, column=1, value=lab).font = Font(name=F, size=10, bold=(lab == "% complete"))
    c = ds.cell(row=r, column=2, value=f_); c.font = Font(name=F, size=10, bold=(lab == "% complete"))
    c.alignment = Alignment(horizontal="center"); c.border = BORD
ds["B9"].number_format = "0%"

ds["A12"] = "EFFORT BY PHASE (solo-days)"; ds["A12"].font = Font(name=F, bold=True, size=11, color=NAVY)
hdr(ds, 13, ["Phase", "Planned", "Done", "% done"], height=18, freeze=False)
phases = ["1 Workshop target", "2 Rail", "3 Plates + UI", "4 Travel kit"]
for i, ph in enumerate(phases):
    r = 14 + i
    ds.cell(row=r, column=1, value=ph).font = Font(name=F, size=10)
    ds.cell(row=r, column=1).border = BORD
    for col, f_ in ((2, f'=SUMIF(Roadmap!$C${r0}:$C${last},$A{r},Roadmap!$G${r0}:$G${last})'),
                    (3, f'=SUMIFS(Roadmap!$G${r0}:$G${last},Roadmap!$C${r0}:$C${last},$A{r},Roadmap!$H${r0}:$H${last},"Done")'),
                    (4, f'=IFERROR(C{r}/B{r},0)')):
        c = ds.cell(row=r, column=col, value=f_); c.font = Font(name=F, size=10)
        c.alignment = Alignment(horizontal="center"); c.border = BORD
    ds.cell(row=r, column=4).number_format = "0%"
r = 18
ds.cell(row=r, column=1, value="TOTAL").font = Font(name=F, size=10, bold=True)
ds.cell(row=r, column=1).border = BORD
for col, f_ in ((2, "=SUM(B14:B17)"), (3, "=SUM(C14:C17)"), (4, "=IFERROR(C18/B18,0)")):
    c = ds.cell(row=r, column=col, value=f_); c.font = Font(name=F, size=10, bold=True)
    c.alignment = Alignment(horizontal="center"); c.border = BORD
ds["D18"].number_format = "0%"

ds["A21"] = "TRANSPORT CHECK (live from the Weight Budget and Packing tabs)"
ds["A21"].font = Font(name=F, bold=True, size=11, color=NAVY)
for i, (lab, f_, fmt) in enumerate([
    ("Projected packed weight (kg)", "='Weight Budget'!C19", "0.0"),
    ("Airline limit (kg)", "=23", "0"),
    ("Weight margin (kg)", "=B23-B22", "0.0"),
    ("Weight verdict", '=IF(B22<=23,IF(B22<=20,"PASS with margin","PASS - tight"),"FAIL - over limit")', "@"),
    ("Case linear sum (cm)", "=Packing!C7+Packing!C8+Packing!C9", "0"),
    ("Airline limit (cm)", "=158", "0"),
    ("Size verdict", '=IF(B26<=158,"PASS","FAIL - over limit")', "@")]):
    r = 22 + i
    ds.cell(row=r, column=1, value=lab).font = Font(name=F, size=10, bold=("verdict" in lab))
    c = ds.cell(row=r, column=2, value=f_); c.font = Font(name=F, size=10, bold=("verdict" in lab))
    c.alignment = Alignment(horizontal="center"); c.border = BORD; c.number_format = fmt
for cellref in ("B25", "B28"):
    ds.conditional_formatting.add(cellref, CellIsRule(operator="containsText", formula=[f'NOT(ISERROR(SEARCH("FAIL",{cellref})))'],
        fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(name=F, size=10, bold=True, color="9C0006")))

ds["A31"] = "GATES"; ds["A31"].font = Font(name=F, bold=True, size=11, color=NAVY)
hdr(ds, 32, ["Gate", "Week", "Test", "Action on failure"], height=18, freeze=False)
gates = [
 ("Adoption", 6, "Does a team member reach for it unprompted, and run it unaided?", "Understand why not before adding rail, sensors or a suitcase to it"),
 ("Plate gate", 10, "Prototype plate RMS <= 25 mm on held-out grid positions", "Do not build four. Ship with stock plates + referee counting"),
 ("Packing test", "20-21", "Full dry-run packout under 23 kg and 158 cm linear", "Decide deliberately: oversize fee, freight, or workshop-only"),
 ("Dry run", 24, "Teammate assembles from case, unaided, under 10 min", "The design is ambiguous. Fix keying and labelling, re-run"),
 ("Reliability tripwire", "any", "Two consecutive sessions needing tools to get running", "Stop adding features. Everything here is downstream of it working"),
]
for i, g in enumerate(gates):
    r = 33 + i
    putrow(ds, r, list(g), wrap=(3, 4), center=(2,), band=(i % 2 == 1))
    ds.row_dimensions[r].height = 30
ds.column_dimensions["C"].width = 50
ds.column_dimensions["D"].width = 54
ds.freeze_panes = None
ds.sheet_view.showGridLines = False

# ============================================================ WEIGHT BUDGET
wt = wb.create_sheet("Weight Budget")
title(wt, "Weight Budget - the constraint that is discovered at the check-in desk",
      "Blue = your input. Replace each estimate with an ACTUAL as parts get built and weighed. Size fails visibly; weight fails invisibly.")
hdr(wt, 4, ["Subsystem", "Item detail", "Est. (kg)", "Actual (kg)", "Using (kg)", "Flies?", "Note"],
    widths=[26, 40, 11, 12, 12, 9, 44], height=30)
items = [
 ("Motor", "GM6020 - 468 g confirmed from manufacturer user guide", 0.47, None, "Yes"),
 ("Head structure", "Hub, 4 arms, plate carriers, balance masses", 1.80, None, "Yes"),
 ("Armor plates", "4 x custom small-format plates with sensors bonded", 1.60, None, "Yes"),
 ("Rotating electronics", "Head MCU, front end, slip ring, wiring", 0.30, None, "Yes"),
 ("Base electronics", "Controller, universal-input 24V PSU, enclosure, E-stop", 1.50, None, "Yes"),
 ("Travel stand", "Legs, column, feet, latches - LARGEST UNCERTAINTY", 3.50, None, "Yes"),
 ("Spares kit", "Piezo modules, spare plate face, latches, fasteners", 0.80, None, "Yes"),
 ("Case + foam", "Shell, cut foam, corner protection - heavier than expected", 4.50, None, "Yes"),
 ("Cables + PSU lead", "Mains lead, DC leads, spare connectors", 0.40, None, "Yes"),
 ("Workshop base", "Ballasted steel pedestal - stays in Singapore", 12.00, None, "No"),
 ("Rail assembly", "Extrusion, carriage, belt, motor - stays in Singapore", 15.00, None, "No"),
]
w0 = 5
for i, (sub, det, est, act, flies) in enumerate(items):
    r = w0 + i
    putrow(wt, r, [sub, det, est, act, f'=IF(D{r}="",C{r},D{r})', flies, ""],
           wrap=(2, 7), center=(3, 4, 5, 6), blue=(4,), band=(i % 2 == 1))
    for col in (3, 4, 5): wt.cell(row=r, column=col).number_format = "0.00"
wl = w0 + len(items) - 1
wt.cell(row=wl+2, column=2, value="SUBTOTAL - items that fly").font = Font(name=F, bold=True, size=10)
c = wt.cell(row=wl+2, column=5, value=f'=SUMIF(F{w0}:F{wl},"Yes",E{w0}:E{wl})')
c.font = Font(name=F, bold=True, size=10); c.number_format = "0.00"; c.alignment = Alignment(horizontal="center"); c.border = BORD
wt.cell(row=wl+3, column=2, value="Contingency (% - appropriate before any part exists)").font = Font(name=F, size=10)
c = wt.cell(row=wl+3, column=4, value=0.25); c.font = Font(name=F, size=10, color=BLUE); c.number_format = "0%"; c.alignment = Alignment(horizontal="center"); c.border = BORD
c = wt.cell(row=wl+3, column=5, value=f"=E{wl+2}*D{wl+3}"); c.font = Font(name=F, size=10); c.number_format = "0.00"; c.alignment = Alignment(horizontal="center"); c.border = BORD
wt.cell(row=wl+4, column=2, value="PROJECTED PACKED WEIGHT").font = Font(name=F, bold=True, size=10)
c = wt.cell(row=wl+4, column=3, value=f"=E{wl+2}+E{wl+3}"); c.font = Font(name=F, bold=True, size=11); c.number_format = "0.00"; c.alignment = Alignment(horizontal="center"); c.border = BORD
wt.cell(row=wl+6, column=2, value="Airline limit (kg)").font = Font(name=F, size=10)
c = wt.cell(row=wl+6, column=3, value=23); c.font = Font(name=F, size=10, color=BLUE); c.number_format = "0"; c.alignment = Alignment(horizontal="center"); c.border = BORD
wt.cell(row=wl+7, column=2, value="Internal target (kg) - margin for a heavier case than planned").font = Font(name=F, size=10)
c = wt.cell(row=wl+7, column=3, value=20); c.font = Font(name=F, size=10, color=BLUE); c.number_format = "0"; c.alignment = Alignment(horizontal="center"); c.border = BORD
wt.cell(row=wl+8, column=2, value="MARGIN vs. airline limit (kg)").font = Font(name=F, bold=True, size=10)
c = wt.cell(row=wl+8, column=3, value=f"=C{wl+6}-C{wl+4}"); c.font = Font(name=F, bold=True, size=10); c.number_format = "0.00"; c.alignment = Alignment(horizontal="center"); c.border = BORD
wt.cell(row=wl+9, column=2, value="VERDICT").font = Font(name=F, bold=True, size=10)
c = wt.cell(row=wl+9, column=3, value=f'=IF(C{wl+4}<=C{wl+7},"PASS with margin",IF(C{wl+4}<=C{wl+6},"PASS - tight","FAIL - over limit"))')
c.font = Font(name=F, bold=True, size=10); c.alignment = Alignment(horizontal="center"); c.border = BORD
wt.cell(row=wl+11, column=1, value="The stand and the case are ~44% of flying mass and are the two least-defined items. Build and weigh both EARLY in Phase 4, not at the end.").font = Font(name=F, size=9, italic=True, color="9C0006")
wt.cell(row=wl+12, column=1, value="Source: GM6020 mass from the RoboMaster GM6020 Brushless DC Motor User Guide. Every other figure is an estimate until an Actual is entered.").font = Font(name=F, size=9, italic=True, color="666666")
wt.sheet_view.showGridLines = False

# ============================================================ PACKING
pk = wb.create_sheet("Packing")
title(pk, "Case Sizing + Packing Manifest",
      "158 cm is a SUM of the three dimensions, not a length. Case shape is a free design variable - choose it AFTER the largest irreducible part is known.")
pk.column_dimensions["A"].width = 34
for col in "BCDEFG": pk.column_dimensions[col].width = 15

pk["A4"] = "CASE DIMENSIONS (external, cm)"; pk["A4"].font = Font(name=F, bold=True, size=11, color=NAVY)
pk["A5"] = "Blue = your input. Set these once the part list is final."; pk["A5"].font = Font(name=F, size=9, italic=True, color="666666")
for i, (lab, val) in enumerate([("Length", 85), ("Width", 45), ("Height", 28)]):
    r = 7 + i
    pk.cell(row=r, column=1, value=lab).font = Font(name=F, size=10)
    pk.cell(row=r, column=1).border = BORD
    c = pk.cell(row=r, column=3, value=val)
    c.font = Font(name=F, size=10, bold=True, color=BLUE); c.fill = PatternFill("solid", fgColor="FFFFCC")
    c.alignment = Alignment(horizontal="center"); c.border = BORD; c.number_format = "0"
for i, (lab, f_, fmt) in enumerate([
    ("LINEAR SUM (L+W+H)", "=C7+C8+C9", "0"),
    ("Airline limit", "=158", "0"),
    ("Margin (cm)", "=C12-C11", "0"),
    ("VERDICT", '=IF(C11<=158,"PASS","FAIL - over limit")', "@"),
    ("Longest internal member (cm, approx.)", "=C7-4", "0"),
    ("Internal diagonal (cm)", "=SQRT((C7-4)^2+(C8-4)^2+(C9-4)^2)", "0")]):
    r = 11 + i
    pk.cell(row=r, column=1, value=lab).font = Font(name=F, size=10, bold=(lab in ("LINEAR SUM (L+W+H)", "VERDICT")))
    pk.cell(row=r, column=1).border = BORD
    c = pk.cell(row=r, column=3, value=f_); c.font = Font(name=F, size=10, bold=(lab in ("LINEAR SUM (L+W+H)", "VERDICT")))
    c.alignment = Alignment(horizontal="center"); c.border = BORD; c.number_format = fmt

pk["A19"] = "CASE SHAPE OPTIONS - all sum to ~158 cm"; pk["A19"].font = Font(name=F, bold=True, size=11, color=NAVY)
hdr(pk, 20, ["Shape (L x W x H, cm)", "Sum", "Longest member", "Comment"], height=18, freeze=False)
pk.column_dimensions["D"].width = 62
for i, (shape, s, lm, note) in enumerate([
    ("70 x 48 x 40", 158, "~66 cm", "Conventional suitcase proportions. Easy to handle, poor length."),
    ("85 x 45 x 28", 158, "~81 cm", "Balanced. Current working assumption above."),
    ("105 x 32 x 20", 157, "~101 cm", "Ski/golf-case proportions. Best length, awkward, more bending risk."),
    ("120 x 24 x 14", 158, "~116 cm", "Extreme. Fragile as a case, hard to protect. Not recommended.")]):
    r = 21 + i
    putrow(pk, r, [shape, s, lm, note], wrap=(4,), center=(1, 2, 3), band=(i % 2 == 1))

pk["A27"] = "PACKING MANIFEST - one row per part. An empty foam pocket is a visible missing part."
pk["A27"].font = Font(name=F, bold=True, size=11, color=NAVY)
hdr(pk, 28, ["Part", "Qty", "Longest dim (cm)", "Mass each (kg)", "Total (kg)", "Foam pocket #", "Packed?"], height=30, freeze=False)
parts = [
 ("Head hub with motor", 1, 22, 1.10, "P1"), ("Head arm + plate carrier", 4, 20, 0.18, "P2-P5"),
 ("Armor plate assembly", 4, 15, 0.40, "P6-P9"), ("Stand column section", 2, 45, 0.60, "P10-P11"),
 ("Stand leg", 3, 50, 0.55, "P12-P14"), ("Stand foot / ballast pocket", 3, 12, 0.10, "P15-P17"),
 ("Electronics enclosure", 1, 25, 1.50, "P18"), ("PSU + mains lead", 1, 20, 0.40, "P19"),
 ("Cable set", 1, 15, 0.20, "P20"), ("Spares kit box", 1, 20, 0.80, "P21"),
 ("Setup card + customs list", 1, 30, 0.05, "lid"),
]
p0 = 29
for i, (nm, qty, dim, mass, pocket) in enumerate(parts):
    r = p0 + i
    putrow(pk, r, [nm, qty, dim, mass, f"=B{r}*D{r}", pocket, "No"], center=(2, 3, 4, 5, 6, 7), blue=(2, 3, 4), band=(i % 2 == 1))
    for col in (4, 5): pk.cell(row=r, column=col).number_format = "0.00"
pl = p0 + len(parts) - 1
pk.cell(row=pl+2, column=4, value="CONTENTS TOTAL (kg)").font = Font(name=F, bold=True, size=10)
c = pk.cell(row=pl+2, column=5, value=f"=SUM(E{p0}:E{pl})"); c.font = Font(name=F, bold=True, size=10); c.number_format = "0.00"; c.alignment = Alignment(horizontal="center"); c.border = BORD
pk.cell(row=pl+3, column=4, value="Longest part (cm)").font = Font(name=F, bold=True, size=10)
c = pk.cell(row=pl+3, column=5, value=f"=MAX(C{p0}:C{pl})"); c.font = Font(name=F, bold=True, size=10); c.alignment = Alignment(horizontal="center"); c.border = BORD
pk.cell(row=pl+4, column=4, value="Fits the case?").font = Font(name=F, bold=True, size=10)
c = pk.cell(row=pl+4, column=5, value=f'=IF(E{pl+3}<=C15,"YES","NO - part is longer than the case interior")')
c.font = Font(name=F, bold=True, size=10); c.alignment = Alignment(horizontal="center"); c.border = BORD
pk.cell(row=pl+5, column=4, value="Items packed").font = Font(name=F, size=10)
c = pk.cell(row=pl+5, column=5, value=f'=COUNTIF(G{p0}:G{pl},"Yes")&" / "&COUNTA(A{p0}:A{pl})'); c.font = Font(name=F, size=10); c.alignment = Alignment(horizontal="center"); c.border = BORD
dvp = DataValidation(type="list", formula1='"No,Yes"', allow_blank=True)
pk.add_data_validation(dvp); dvp.add(f"G{p0}:G{pl}")
pk.conditional_formatting.add(f"G{p0}:G{pl}", CellIsRule(operator="equal", formula=['"Yes"'],
    fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(name=F, size=10, color="006100")))
pk.cell(row=pl+7, column=1, value="Part dimensions and masses are placeholders until real parts exist. Update them as you build - this tab is the single source of truth for the Dashboard transport check.").font = Font(name=F, size=9, italic=True, color="666666")
pk.sheet_view.showGridLines = False

# ============================================================ BOM
bs = wb.create_sheet("BOM")
title(bs, "Bill of Materials", "Blue = your input. Phase column lets you stage funding: Phases 1-2 deliver the training tool, Phases 3-4 add measurement and portability.")
hdr(bs, 4, ["Group", "Item", "Phase", "Qty", "Unit $", "Line $", "Own?"], widths=[22, 44, 8, 6, 10, 10, 10], height=26)
bom = [
 ("Head", "GM6020 gimbal motor (468 g, 1.2 Nm, 18 mm hollow shaft)", 1, 1, 110, "Likely"),
 ("Head", "Hub + arm stock, plate carriers, balance masses", 1, 1, 80, "No"),
 ("Head", "Hardware: captive fasteners, latches, bearings", 1, 1, 20, "No"),
 ("Workshop base", "Steel plate, ballast, mounting hardware", 1, 1, 40, "Partly"),
 ("Rail", "Aluminium extrusion, 2.5-3 m", 2, 1, 90, "No"),
 ("Rail", "V-wheel carriage, GT belt, pulleys, idlers", 2, 1, 55, "No"),
 ("Rail", "Drive motor + limit switches", 2, 1, 45, "Maybe"),
 ("Plates", "4 x plate substrate + replaceable impact faces", 3, 4, 22, "No"),
 ("Plates", "Piezo elements", 3, 16, 1.5, "No"),
 ("Plates", "Analogue front end: comparators, op-amps, PCB", 3, 4, 9.5, "No"),
 ("Rotating elec.", "Head MCU board", 3, 1, 20, "No"),
 ("Rotating elec.", "Slip ring capsule, hollow-shaft compatible", 3, 1, 35, "No"),
 ("Rotating elec.", "Connectors, wiring, strain relief", 3, 1, 25, "No"),
 ("Base elec.", "Base controller board", 1, 1, 30, "Maybe"),
 ("Base elec.", "Universal-input 24 V PSU (100-240 V)", 1, 1, 45, "No"),
 ("Base elec.", "Enclosure, hardware E-stop, wiring", 1, 1, 35, "No"),
 ("Travel stand", "Extrusion or tube stock, feet, ballast pockets", 4, 1, 60, "No"),
 ("Travel stand", "Draw latches, cam locks, captive hardware", 4, 1, 40, "No"),
 ("Case", "Case shell", 4, 1, 55, "No"),
 ("Case", "Custom-cut foam, corner protection", 4, 1, 35, "No"),
 ("Spares", "Piezo modules, spare plate face, latches, fasteners", 4, 1, 50, "No"),
]
b0 = 5
for i, (grp, item, ph, qty, unit, own) in enumerate(bom):
    r = b0 + i
    putrow(bs, r, [grp, item, ph, qty, unit, f"=D{r}*E{r}", own], wrap=(2,), center=(3, 4, 5, 6, 7), blue=(4, 5), band=(i % 2 == 1))
    for col in (5, 6): bs.cell(row=r, column=col).number_format = '$#,##0'
blast = b0 + len(bom) - 1
rows = [("SUBTOTAL (all phases)", f"=SUM(F{b0}:F{blast})", True),
        ("Less items likely already owned", f'=SUMIF(G{b0}:G{blast},"Likely",F{b0}:F{blast})', False),
        ("LIKELY SPEND", f"=F{blast+2}-F{blast+3}", True),
        ("", "", False),
        ("Phase 1 (workshop target)", f"=SUMIF(C{b0}:C{blast},1,F{b0}:F{blast})", False),
        ("Phase 2 (rail)", f"=SUMIF(C{b0}:C{blast},2,F{b0}:F{blast})", False),
        ("Phase 3 (plates + electronics)", f"=SUMIF(C{b0}:C{blast},3,F{b0}:F{blast})", False),
        ("Phase 4 (travel kit)", f"=SUMIF(C{b0}:C{blast},4,F{b0}:F{blast})", False),
        ("", "", False),
        ("Phases 1-2 only (the training tool)", f"=F{blast+6}+F{blast+7}", True),
        ("Phases 3-4 (measurement + portability)", f"=F{blast+8}+F{blast+9}", True)]
for i, (lab, f_, bold) in enumerate(rows):
    r = blast + 2 + i
    if not lab: continue
    bs.cell(row=r, column=5, value=lab).font = Font(name=F, bold=bold, size=10)
    c = bs.cell(row=r, column=6, value=f_); c.font = Font(name=F, bold=bold, size=10); c.number_format = '$#,##0'; c.border = BORD
bs.cell(row=blast+15, column=1, value="Staging option: fund Phases 1-2 now ($550, the training tool), decide Phases 3-4 at the Week 8 review - which also defers plate spend past the Week 10 gate.").font = Font(name=F, size=9, italic=True, color="9C0006")
bs.sheet_view.showGridLines = False

# ============================================================ RISKS
rs = wb.create_sheet("Risks")
title(rs, "Risk Register", "Score = Impact x Likelihood (1-5 each). Blue = your scoring inputs. Review at every gate.")
hdr(rs, 4, ["ID", "Risk", "Imp", "Lik", "Score", "Severity", "Mitigation", "Review at"],
    widths=[6, 40, 7, 7, 8, 11, 58, 12], height=30)
risks = [
 ("R1","Slotted joints loosen under vibration and repeated impact",5,4,"REQ-M9: separate locating from clamping in every joint. Endurance run and full joint re-check in Phase 1, not Phase 4.","Wk 5"),
 ("R2","Only the builder can operate or fix it",5,4,"M2 makes this a tested requirement, not an aspiration. Keying, labelling, self-test and a setup card in the case. A teammate runs Phase 1 acceptance.","Wk 6"),
 ("R3","Weight creep pushes the packed kit past 23 kg",4,4,"Weigh every part as it is made. Stand and case are 44% of flying mass and least defined - build and weigh them EARLY in Phase 4.","Wk 20"),
 ("R4","Piezo bonds fatigue and silently corrupt localisation",4,4,"REQ-E6 per-channel drift detection; field-replaceable modules; boot self-test on the Health screen.","Wk 10"),
 ("R5","Solo builder, in-season workload, project stalls part-built",4,4,"Phase 1 delivers standalone value in 6 weeks; every later phase is independently useful. Stopping at any phase is acceptable.","Wk 6"),
 ("R6","Travel stand cannot be both packable and stable enough",4,3,"Contained by the architecture - workshop use is unaffected. Ballast filled on site rather than chasing stiffness in the suitcase.","Wk 18"),
 ("R7","No usable mains power at the competition venue",4,3,"Dual DC input accepting a borrowed RM pack. No battery can be flown regardless. Confirm venue power (Q3).","Wk 8"),
 ("R8","Case fails baggage handling; parts arrive damaged",4,3,"Drop test before travelling. Spares kit inside the case. Irreplaceable electronics in hand luggage.","Wk 22"),
 ("R9","GM6020 torque marginal once the head is fully built",3,3,"Budget head inertia before building. Keep plate radius small - inertia scales with r^2, the cheapest lever available.","Wk 4"),
 ("R10","Injury from the spinning head or projectiles",5,2,"Hardware E-stop cutting motor power directly. Balance verification before first full-speed run. Captive fasteners. Eye protection.","Wk 4"),
 ("R11","Slip ring wears or becomes intermittent",3,3,"REQ-E1 keeps only power and digital across the interface. REQ-E3 keeps a wireless data path open as a fallback.","Wk 11"),
 ("R12","Referee gear unusable on a non-robot fixture at a venue",2,3,"Custom plates work standalone, so referee is a workshop cross-check not a dependency. Confirm the rules position (Q4).","Wk 8"),
]
q0 = 5
for i, (rid, risk, imp, lik, mit, rev) in enumerate(risks):
    r = q0 + i
    putrow(rs, r, [rid, risk, imp, lik, f"=C{r}*D{r}",
                   f'=IF(E{r}>=15,"Critical",IF(E{r}>=9,"High",IF(E{r}>=4,"Medium","Low")))', mit, rev],
           wrap=(2, 7), center=(1, 3, 4, 5, 6, 8), blue=(3, 4), bold=(1,), band=(i % 2 == 1))
rl = q0 + len(risks) - 1
for val, fill, col in (("Critical","FFC7CE","9C0006"), ("High","FFD9CC","C55A11"), ("Medium","FFEB9C","9C6500"), ("Low","C6EFCE","006100")):
    rs.conditional_formatting.add(f"F{q0}:F{rl}", CellIsRule(operator="equal", formula=[f'"{val}"'],
        fill=PatternFill("solid", fgColor=fill), font=Font(name=F, size=10, bold=True, color=col)))
rs.cell(row=rl+2, column=2, value="Highest score:").font = Font(name=F, bold=True, size=10)
c = rs.cell(row=rl+2, column=5, value=f"=MAX(E{q0}:E{rl})"); c.font = Font(name=F, bold=True, size=10); c.alignment = Alignment(horizontal="center")
rs.cell(row=rl+3, column=2, value="Thresholds: Critical >=15, High >=9, Medium >=4, Low <4.").font = Font(name=F, size=9, italic=True, color="666666")
rs.sheet_view.showGridLines = False

# ============================================================ METRICS
ms = wb.create_sheet("Metrics")
title(ms, "Success Metrics", "Blue = fill in as measurements land.")
hdr(ms, 4, ["ID", "Metric", "Target", "Actual", "Due", "Status", "Why it is the right measure"],
    widths=[6, 30, 28, 13, 8, 12, 48], height=30)
mets = [
 ("M1","Setup time, case to first shot","<= 10 min, one person, no tools","Wk 24","Decides whether it actually gets used"),
 ("M2","Use by someone who did not build it","Team member runs a full session unaided, first attempt","Wk 6","The primary user test. Tests the design, not the builder's memory"),
 ("M3","Sessions between tool-requiring failures",">= 20","ongoing","Encodes 'switch it on and start using it'"),
 ("M4","Packed size","<= 158 cm linear, case included","Wk 21","Binary pass/fail at the check-in desk"),
 ("M5","Packed weight","<= 23 kg, target <= 20 kg","Wk 21","3 kg margin absorbs spares and a heavier case"),
 ("M6","Spin speed consistency","<= 2% from commanded, 0.25-3.0 rev/s","Wk 4","Two sessions a week apart must be comparable"),
 ("M7","Hit localisation accuracy","<= 25 mm RMS, >= 95% quadrant","Wk 17","Honest limit for piezo TDOA on a small plate"),
 ("M8","Adoption",">= 8 team sessions in the first month, by >= 2 different people","Wk 10","A tool one person uses is a rig; a tool the team uses is equipment"),
]
m0 = 5
for i, (mid, met, tgt, due, why) in enumerate(mets):
    r = m0 + i
    putrow(ms, r, [mid, met, tgt, "", due, "Pending", why], wrap=(2, 3, 7), center=(1, 4, 5, 6), blue=(4,), bold=(1,), band=(i % 2 == 1))
ml = m0 + len(mets) - 1
dvm = DataValidation(type="list", formula1='"Pending,On track,At risk,Met,Missed,Descoped"', allow_blank=True)
ms.add_data_validation(dvm); dvm.add(f"F{m0}:F{ml}")
for val, fill, col in (("Met","C6EFCE","006100"), ("Missed","FFC7CE","9C0006"), ("At risk","FFEB9C","9C6500")):
    ms.conditional_formatting.add(f"F{m0}:F{ml}", CellIsRule(operator="equal", formula=[f'"{val}"'],
        fill=PatternFill("solid", fgColor=fill), font=Font(name=F, size=10, color=col)))
ms.cell(row=ml+2, column=2, value="M2 and M8 are the pair that matter. Every other metric can be satisfied by a machine only its builder can operate.").font = Font(name=F, size=9, italic=True, color="9C0006")
ms.sheet_view.showGridLines = False

# ============================================================ OPEN QUESTIONS
qs = wb.create_sheet("Open Questions")
title(qs, "Open Questions", "Q1 and Q7 gate Week 1-2. Q1 can reorder the whole back half of the plan.")
hdr(qs, 4, ["ID", "Question", "Blocks", "Needed by", "Status", "Answer"],
    widths=[6, 54, 36, 11, 12, 32], height=28)
qq = [
 ("Q1","Which competition and date is the travel deadline? RMUL 2026 mainland sites ran Mar-May 2026, so this plan assumes the 2027 cycle (~Mar 2027). If the real date is earlier, Phase 4 must move ahead of Phase 3.","All Phase 4 scheduling","Wk 2"),
 ("Q7","Do you have 3D printing and machining access, or is everything cut by hand?","Realistic effort for every mechanical part","Wk 1"),
 ("Q6","What plate height above the floor matches the robots you most need to practise against?","Stand height, therefore packed length","Wk 3"),
 ("Q2","Is the ~$830 budget approvable, or should it be staged at the Week 8 review?","Phase 3 and 4 procurement","Wk 4"),
 ("Q5","How long is the workshop rail, and what mounting does the space allow?","Phase 2 rail design","Wk 5"),
 ("Q3","Is mains power available near a usable test area in a competition pit?","Power architecture and on-site fallback","Wk 8"),
 ("Q4","Can referee gear be used on a non-robot fixture at a venue, and is a spare set available?","Whether referee cross-check extends beyond the workshop","Wk 8"),
]
qq0 = 5
for i, (qid, q, blk, due) in enumerate(qq):
    r = qq0 + i
    putrow(qs, r, [qid, q, blk, due, "Open", ""], wrap=(2, 3, 6), center=(1, 4, 5), blue=(6,), bold=(1,), band=(i % 2 == 1))
    qs.row_dimensions[r].height = 40
qsl = qq0 + len(qq) - 1
dvq = DataValidation(type="list", formula1='"Open,Answered,Blocked,Moot"', allow_blank=True)
qs.add_data_validation(dvq); dvq.add(f"E{qq0}:E{qsl}")
for val, fill, col in (("Answered","C6EFCE","006100"), ("Open","FFEB9C","9C6500")):
    qs.conditional_formatting.add(f"E{qq0}:E{qsl}", CellIsRule(operator="equal", formula=[f'"{val}"'],
        fill=PatternFill("solid", fgColor=fill), font=Font(name=F, size=10, color=col)))
qs.sheet_view.showGridLines = False

wb.calculation.fullCalcOnLoad = True
wb.save("/sessions/great-vigilant-edison/mnt/outputs/Portable-Aimbot-Target-Roadmap.xlsx")
print("saved")
